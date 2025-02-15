import openpyxl

workbook = openpyxl.load_workbook('file.xlsx')
sheet = workbook.active

# just want rows with this keyword:
keyword = 'Singulator 1'

rows_to_keep = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        if keyword.lower() in str(cell.value).lower():
            rows_to_keep.append(row[0].row)

for row_num in range(sheet.max_row, 0, -1):
    if row_num not in rows_to_keep:
        sheet.delete_rows(row_num)

workbook.save('file.xlsx')
